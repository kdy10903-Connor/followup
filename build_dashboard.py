#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from datetime import datetime, timedelta
import html
import os

excel_file = r"C:\Users\kdy0613\Downloads\2026년 전사 팔로업 업무_r2 (1).xlsx"
output_file = r"C:\Users\kdy0613\Desktop\경전실\26년 전사 팔로업 업무 관리\followup\index.html"

# 엑셀 파일 읽기
try:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if "Dashboard" not in wb.sheetnames:
        print("Error: Dashboard 시트를 찾을 수 없습니다.")
        exit(1)

    ws = wb["Dashboard"]

    # 데이터 파싱
    kpi_row = 2  # KPI 행
    kpi_values = [ws.cell(kpi_row, col).value for col in range(1, 6)]

    # 조직별 색상 맵
    org_colors = {
        "경영전략실": "#FF6B6B",
        "RPM 추진실": "#4ECDC4",
        "SK 추진실": "#45B7D1",
        "CTO 조직": "#96CEB4",
        "경영조직": "#FFEAA7",
        "신사옥 추진실": "#DDA15E",
        "중국법인": "#BC6C25",
    }

    # HTML 생성
    html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>2026년 전사 팔로업 업무 대시보드</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .header {{ text-align: center; margin-bottom: 40px; }}
        .header h1 {{ font-size: 28px; color: #333; margin-bottom: 10px; }}
        .header p {{ color: #999; font-size: 14px; }}
        .kpi-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 15px; margin-bottom: 40px; }}
        .kpi-card {{ background: white; padding: 20px; border-radius: 8px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .kpi-card .label {{ font-size: 12px; color: #999; margin-bottom: 8px; text-transform: uppercase; }}
        .kpi-card .value {{ font-size: 32px; font-weight: bold; color: #333; }}
        .section {{ background: white; border-radius: 8px; overflow: hidden; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        .section-header {{ background: #f8f8f8; padding: 16px; border-left: 4px solid #4a90e2; font-size: 16px; font-weight: 600; color: #333; }}
        .section-content {{ padding: 16px; }}
        .item {{ display: flex; justify-content: space-between; align-items: center; padding: 12px 0; border-bottom: 1px solid #eee; }}
        .item:last-child {{ border-bottom: none; }}
        .item-title {{ flex: 1; }}
        .org-tag {{ display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; color: white; margin-right: 8px; }}
        .d-day {{ display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: bold; }}
        .critical {{ background: #FF6B6B; color: white; }}
        .urgent {{ background: #FFB347; color: white; }}
        .warn {{ background: #FFD93D; color: #333; }}
        .done {{ background: #6BCB77; color: white; }}
        .footer {{ text-align: center; color: #999; font-size: 12px; margin-top: 40px; }}
        .updated {{ color: #666; font-size: 12px; margin-top: 10px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>2026년 전사 팔로업 업무 대시보드</h1>
            <p>주간 팔로업 현황</p>
            <div class="updated">마지막 업데이트: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M')}</div>
        </div>

        <div class="kpi-grid">
"""

    # KPI 카드
    kpi_labels = ["총 항목", "7일 내", "8일 이후", "미정", "완료"]
    for label, value in zip(kpi_labels, kpi_values):
        html_content += f"""            <div class="kpi-card">
                <div class="label">{label}</div>
                <div class="value">{value or 0}</div>
            </div>
"""

    html_content += """        </div>

        <div class="section">
            <div class="section-header">7일 내 팔로업 사항</div>
            <div class="section-content">
                <p style="color: #999; font-size: 12px;">데이터를 로드 중입니다...</p>
            </div>
        </div>

        <div class="section">
            <div class="section-header">8일 이후 팔로업 사항</div>
            <div class="section-content">
                <p style="color: #999; font-size: 12px;">데이터를 로드 중입니다...</p>
            </div>
        </div>

        <div class="footer">
            <p>마이다스 전사 팔로업 업무 관리 시스템</p>
        </div>
    </div>
</body>
</html>
"""

    # 파일 저장
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f"✅ 대시보드가 생성되었습니다: {output_file}")
    print(f"KPI 값: {kpi_values}")

except Exception as e:
    print(f"❌ 오류: {e}")
    import traceback
    traceback.print_exc()
